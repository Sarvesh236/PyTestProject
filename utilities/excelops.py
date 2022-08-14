import openpyxl


class ExcelOperations:
    def __init__(self, file_path, sheet_name):
        global book
        global sheet
        book = openpyxl.load_workbook(file_path)
        sheet = book[sheet_name]

    def fetch_rows(self):
        return sheet.max_row

    def fetch_columns(self):
        return sheet.max_column

    def fetch_testdata(self,row,col):
        return sheet.cell(row=row, column=col).value
