import openpyxl

class HomePageData:

    test_homepage_data = [{"name": "Sarvesh", "email": "sarvesh@gmail.com", "password": "1234", "gender": "Male"},
                            {"name": "Taksha", "email": "taksha@gmail.com", "password": "7890", "gender": "Female"}]


    @staticmethod
    def getTestData(test_case):

        book = openpyxl.load_workbook(
            "C:\\Users\\sarve\\PycharmProjects\\selfFrameworkProject\\TestData\\PythonDemo.xlsx")
        sheet = book.active
        Dict = {}

        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [Dict]
