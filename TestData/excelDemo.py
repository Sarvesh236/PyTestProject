import openpyxl

book = openpyxl.load_workbook("C:\\Users\\sarve\\PycharmProjects\\selfFrameworkProject\\TestData\\PythonDemo.xlsx")

# sheets = book.sheetnames
# sheet = book[sheets[1]]                      -- To select diffrent sheets
# cell = sheet.cell(row=1, column=2)
# print(cell.value)

sheet = book.active
# cell = sheet.cell(row=1, column=2)
# print(cell.value)

# sheet.cell(row=2, column=2).value = "Sarvesh"
# print(sheet.cell(row=2, column=2).value)
# print(sheet.max_row)
# print(sheet.max_column)
# print(sheet['A7'].value)

details = []
for i in range(2, sheet.max_row+1):
    Dict = {}
    for j in range(2, sheet.max_column+1):
        Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
    details.append(Dict)

print(details)
